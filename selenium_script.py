import os
import shutil
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
import re
import time
# import HibbettOpenpyexcel

#INPUT INTO THIS LIST TO RUN SCRIPT PROPERLY

site_numbers = ['']


#TODO:make script to fill in required fields and navigate webpage = DONE
#TODO:make code that will loop through a list of completed sites and their kit numbers = Waiting for site list
#TODO:make code to loop through completed test results and zip them for uploading = Need site list to corrosponds with
# site number

site_regex = re.compile(r'(?m)^(\d+).*')

files = os.listdir('C:\hibbett-test-results')

filenames = []

for item in files:
    filenames.append(item)

print(filenames)

Path = "C:\\Program Files (x86)\\chromedriver.exe"
driver = webdriver.Chrome('C:\Program Files (x86)\chromedriver.exe')

wb = openpyxl.load_workbook('''C:\hibbett-test-results\site-list\Wachter - Hibbett 907 Site List.xlsx''')

get_sheet_name = 'Hibbett-City Gear Store List'
sheet = wb.get_sheet_by_name(get_sheet_name)



def get_filename():

    for item in filenames:
        if file_regex.search(item) != None:
            return f"{item}"
        else:
            pass
def get_filename2():

    for item in filenames:
        if file_regex2.search(item) != None:
            return f"{item}"
        else:
            pass
site_deet_list = {}
def get_site_deets(site_number):

    site_regex = re.compile(f"{site_number}")
    for num in range(2, sheet.max_row):
        if site_regex.search(f"{sheet[f'A{num}'].value}") != None:
            site_deet_list['address']= f"{sheet[f'L{num}'].value}"
            site_deet_list['city'] = f"{sheet[f'N{num}'].value}"
            site_deet_list['state'] = f"{sheet[f'O{num}'].value}"
            site_deet_list['zip'] = f"{sheet[f'P{num}'].value}"
            site_deet_list['site_num'] = f"{sheet[f'A{num}'].value}"
        else:
            pass




driver.get("https://hubbellwiringsystems.com/installercorner/login/")

email = driver.find_element_by_id("username-79")

password =  driver.find_element_by_id("user_password-79")

email.clear()
email.send_keys("REDACTED)

password.clear()
password.send_keys("REDACTED")#will be a variable read from a json file for security
password.send_keys(Keys.ENTER)


for num in site_numbers:
# for file in files:
#     filenames.append(file)
#     site_num = ""
#     set_filenames = set(filenames)
#     for files in set_filenames:
#         if site_regex.search(files) != None:
#             site_num = site_regex.search(files).group(1)
#         else:
#             pass
    file_regex = re.compile(f"(^{num} -- Hibbitt Sports Test -- [0-9]{{1,2}}-[0-9]{{1,2}}-2021\.zip$)|(^{num} -- Hibbitt Sports -- [0-9]{{1,2}}-[0-9]{{1,2}}-2021\.zip$)")
    file_regex2 = re.compile(f"(^{num} -- Hibbitt Sports Test -- [0-9]{{1,2}}-[0-9]{{1,2}}-2021\.pdf$)|(^{num} -- Hibbitt Sports -- [0-9]{{1,2}}-[0-9]{{1,2}}-2021\.pdf$)")
    get_site_deets(num)
# excel_path = ""
#
# wb = openpyxl.load_workbook(excel_path)
#
# sheet = wb.active


    # # while True:
    #
    #
    #
    driver.get('https://hubbellwiringsystems.com/installercorner/submit-a-claim/')

        # STEP 1 OF 4: STRUCTURED CABLING SYSTEM REGISTRATION REQUEST FORM

    project_name = driver.find_element_by_id("input_1_6")
    project_name.send_keys("Hibbett Sports")

    hubbell_contact_email = driver.find_element_by_id("input_1_81")
    hubbell_contact_email.send_keys("dkreb@hubbell.com")

        #Certified Installation Comapany Infomation

    company = driver.find_element_by_id("input_1_8")
    company.clear()
    company.send_keys("Wachter Inc.")

    str_addr = driver.find_element_by_id("input_1_14_1")
    str_addr.clear()
    str_addr.send_keys("1419 W Monroe Ave")

    city = driver.find_element_by_id("input_1_14_3")
    city.clear()
    city.send_keys("Lowell")

    state = driver.find_element_by_id("input_1_14_4")
    state.clear()
    state.send_keys("Arkansas")

    ZIP = driver.find_element_by_id("input_1_14_5")
    ZIP.clear()
    ZIP.send_keys("72745")

    country = Select(driver.find_element_by_id("input_1_14_6"))
    country.select_by_visible_text("United States")

    frst_name = driver.find_element_by_id("input_1_15_3")
    frst_name.clear()
    frst_name.send_keys("James")

    lst_name = driver.find_element_by_id("input_1_15_6")
    lst_name.clear()
    lst_name.send_keys("Moore")

    cntct_num = driver.find_element_by_id("input_1_67")
    cntct_num.clear()
    cntct_num.send_keys("(479) 361-7913")

    cntct_email = driver.find_element_by_id("input_1_16")
    cntct_email.clear()
    cntct_email.send_keys("james.moore@wachter.com")

    sys_design = driver.find_element_by_id("input_1_17")
    sys_design.clear()
    sys_design.send_keys("Debbi Kreb")

    installer = driver.find_element_by_id("input_1_18")
    installer.clear()
    installer.send_keys("Wachter Inc.")

    tester = driver.find_element_by_id("input_1_19")
    tester.clear()
    tester.send_keys("Lantek IV")

        #WARRANTY INFORMATION

    cable_name = driver.find_element_by_id("input_1_22")
    cable_name.clear()
    cable_name.send_keys("NEXTSPEED LINK6 CMP")

    cable_type = driver.find_element_by_id("choice_1_25_2")
    cable_type.click()

    cable_warranty = driver.find_element_by_id("choice_1_24_1")
    cable_warranty.click()

    wrrnty_consent = driver.find_element_by_id("input_1_30_1")
    wrrnty_consent.click()


        #PROJECT INFORMATION

    project_name_2 = driver.find_element_by_id("input_1_32")
    project_name_2.clear()
    project_name_2.send_keys("Hibbett Sports")

    prjct_comp = driver.find_element_by_id("input_1_33")
    prjct_comp.clear()
    prjct_comp.send_keys("Hibbett Sports, Inc.")

    prjct_address = driver.find_element_by_id("input_1_34_1")
    prjct_address.clear()
    prjct_address.send_keys(f"{site_deet_list['address']}")

    prjct_city = driver.find_element_by_id("input_1_34_3")
    prjct_city.clear()
    prjct_city.send_keys(f"{site_deet_list['city']}")

    prjct_state = driver.find_element_by_id("input_1_34_4")
    prjct_state.clear()
    prjct_state.send_keys(f"{site_deet_list['state']}")

    prjct_zip = driver.find_element_by_id("input_1_34_5")
    prjct_zip.clear()
    prjct_zip.send_keys(f"{site_deet_list['zip']}")

    prjct_country = Select(driver.find_element_by_id("input_1_34_6"))
    prjct_country.select_by_visible_text("United States")

    cntct_first_name = driver.find_element_by_id("input_1_35_3")
    cntct_first_name.clear()
    cntct_first_name.send_keys("Justin")

    cntct_last_name = driver.find_element_by_id("input_1_35_6")
    cntct_last_name.clear()
    cntct_last_name.send_keys("Farr")

    cntct_phone = driver.find_element_by_id("input_1_36")
    cntct_phone.clear()
    cntct_phone.send_keys("2053807199")

    prjct_email = driver.find_element_by_id("input_1_37")
    prjct_email.clear()
    prjct_email.send_keys("Justin.Farr@hibbett.com")

    prt_wrk_area1 = driver.find_element_by_xpath('//*[@id="field_1_41"]/div/table/tbody/tr/td[1]/input')
    prt_wrk_area1.clear()
    prt_wrk_area1.send_keys("HSNCUKIT2")

    prt_wrk_area2 = driver.find_element_by_xpath('//*[@id="field_1_41"]/div/table/tbody/tr/td[2]/input')
    prt_wrk_area2.clear()
    prt_wrk_area2.send_keys('1')

    prt_crss_connect1 = driver.find_element_by_xpath('//*[@id="field_1_42"]/div/table/tbody/tr/td[1]/input')
    prt_crss_connect1.clear()
    prt_crss_connect1.send_keys('HSNCUKIT2')

    prt_crss_connect2 = driver.find_element_by_xpath('//*[@id="field_1_42"]/div/table/tbody/tr/td[2]/input')
    prt_crss_connect2.clear()
    prt_crss_connect2.send_keys('1')

    prt_cable1 = driver.find_element_by_xpath('//*[@id="field_1_43"]/div/table/tbody/tr/td[1]/input')
    prt_cable1.clear()
    prt_cable1.send_keys('HSNCUKIT2')

    prt_cable2 = driver.find_element_by_xpath('//*[@id="field_1_43"]/div/table/tbody/tr/td[2]/input')
    prt_cable2.clear()
    prt_cable2.send_keys('1')

    cable_distro = driver.find_element_by_id('input_1_44')
    cable_distro.clear()
    cable_distro.send_keys('W.W. Grainger')

    distro_addr = driver.find_element_by_id('input_1_54_1')
    distro_addr.clear()
    distro_addr.send_keys('222 W Merchandise Mart Plaza')

    distro_city = driver.find_element_by_id('input_1_54_3')
    distro_city.clear()
    distro_city.send_keys('Chicago')

    distro_state = driver.find_element_by_id('input_1_54_4')
    distro_state.clear()
    distro_state.send_keys('IL')

    page1_nxt_button = driver.find_element_by_id('gform_next_button_1_46')
    page1_nxt_button.click()

    #PAGE2:
    driver.implicitly_wait(15)
    time.sleep(3)
    typ_wrk_con = driver.find_element_by_xpath('//*[@id="field_1_56"]/div/table/tbody/tr/td[1]/input')
    typ_wrk_con.click()
    typ_wrk_con.send_keys('HC6RPEB')

    cable_mfg = driver.find_element_by_xpath('//*[@id="field_1_56"]/div/table/tbody/tr/td[2]/input')
    cable_mfg.send_keys('HC6RPEB')

    typ_crss_comp = driver.find_element_by_xpath('//*[@id="field_1_56"]/div/table/tbody/tr/td[3]/input')
    typ_crss_comp.send_keys('HJU6R')

    num_wrk_areas = driver.find_element_by_id('input_1_58')
    num_wrk_areas.clear()
    num_wrk_areas.send_keys('1')

    page2_nxt = driver.find_element_by_id('gform_next_button_1_49')
    page2_nxt.click()

    #PAGE 3
    time.sleep(3)
    page3_nxt = driver.find_element_by_id('gform_next_button_1_63')
    page3_nxt.click()

    #PAGE 4

    time.sleep(3)
    file_input = driver.find_element_by_id('input_1_85')
    file_input.send_keys(f'C:\\hibbett-test-results\\{get_filename()}')

    submit_button = driver.find_element_by_id('gform_submit_button_1')
    submit_button.click()

    time.sleep(30)


    shutil.move(src=f'C:\\hibbett-test-results\\{get_filename()}',
                dst= f'C:\\hibbett-test-results\\completed\\{get_filename()}')
    shutil.move(src=f'C:\\hibbett-test-results\\{get_filename2()}',
                dst= f'C:\\hibbett-test-results\\completed\\{get_filename2()}')

    print(f'submissions for {num}')

print("submissions done")
